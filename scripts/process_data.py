"""
process_data.py — Parses all raw data files and write JSON to public/json_files
Run from the docs/ directory:  python scripts/process_data.py
"""

import csv
import json
import pathlib
from collections import defaultdict

import openpyxl

#Sets up working dirs
ROOT = pathlib.Path(__file__).parent.parent          
DATA = ROOT / "data"             
OUT  = ROOT / "docs/public/json_files"                           
OUT.mkdir(parents=True, exist_ok=True)


# ── CO₂ (NOAA Mauna Loa) ─────────────────────────────────────────────────────
# File: co2_mm_mlo.csv
# Columns: year, month, decimal_date, average, deseasonalized, ndays, sdev, unc
# Comment lines start with '#'.  Sentinel value -9.99 means missing.
# Output: { id, unit, points: [{d:"YYYY-MM", v:avg_ppm, t:deseas_ppm}] }

def process_co2():
    points = []
    with open(DATA / "co2_mm_mlo.csv") as f:
        for raw in f:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            cols = line.split(",")
            if len(cols) < 5:
                continue
            try:
                year  = int(cols[0])
                month = int(cols[1])
                avg   = float(cols[3])
                deseas = float(cols[4])
            except ValueError:
                continue  # header row
            if avg <= 0:
                continue  # -9.99 sentinel
            point = {"d": f"{year}-{month:02d}", "v": round(avg, 2)}
            if deseas > 0:
                point["t"] = round(deseas, 2)
            points.append(point)

    if len(points) < 100:
        print(f"  co2: only {len(points)} points — skipping")
        return
    _write("co2.json", {"id": "co2", "unit": "ppm", "points": points})
    print(f"  co2: {len(points)} monthly points")


# ── GISTEMP (NASA surface temperature) ───────────────────────────────────────
# File: gistemp_glb.csv  (wide format, rows = years, cols = months + J-D annual)
# Missing values encoded as "***".
# Output: { id, unit, points: [{d:"YYYY-MM", v:anomaly, t:annual_mean}] }

def process_gistemp():
    MONTHS = ["Jan","Feb","Mar","Apr","May","Jun",
              "Jul","Aug","Sep","Oct","Nov","Dec"]
    with open(DATA / "gistemp_glb.csv") as f:
        lines = f.readlines()

    hi = next((i for i, l in enumerate(lines) if l.startswith("Year,")), -1)
    if hi < 0:
        print("  gistemp: header row not found — skipping")
        return

    header     = lines[hi].strip().split(",")
    jd_col     = next((i for i, h in enumerate(header) if h.strip() == "J-D"), -1)
    month_cols = [next((i for i, h in enumerate(header) if h.strip() == m), -1)
                  for m in MONTHS]

    points = []
    for line in lines[hi + 1:]:
        cols = line.strip().split(",")
        if not cols[0].strip():
            continue
        try:
            year = int(cols[0])
        except ValueError:
            continue

        month_vals = []
        for col in month_cols:
            s = cols[col].strip() if 0 <= col < len(cols) else ""
            if not s or s == "***":
                month_vals.append(None)
            else:
                try:
                    month_vals.append(float(s))
                except ValueError:
                    month_vals.append(None)

        # t = J-D annual mean for complete years; partial-year mean otherwise
        t_val = None
        if 0 <= jd_col < len(cols):
            s = cols[jd_col].strip()
            if s and s != "***":
                try:
                    t_val = float(s)
                except ValueError:
                    pass
        if t_val is None:
            avail = [v for v in month_vals if v is not None]
            if avail:
                t_val = round(sum(avail) / len(avail), 4)

        for mi, v in enumerate(month_vals):
            if v is None:
                continue
            point = {"d": f"{year}-{mi + 1:02d}", "v": v}
            if t_val is not None:
                point["t"] = t_val
            points.append(point)

    if len(points) < 100:
        print(f"  gistemp: only {len(points)} points — skipping")
        return
    _write("gistemp.json", {"id": "gistemp", "unit": "\u00b0C vs 1951\u20131980", "points": points})
    print(f"  gistemp: {len(points)} monthly points")


# ── Global CO₂ emissions (OWID / GCP) ────────────────────────────────────────
# File: owid_co2_emissions.csv
# Columns: entity, code, year, emissions_total (tonnes CO₂)
# Keeps entity == "World"; converts tonnes → Gt.
# Output: { id, unit, points: [{d:"YYYY", v:Gt}] }

def process_emissions():
    points = []
    with open(DATA / "owid_co2_emissions.csv", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row.get("entity", "").strip() != "World":
                continue
            try:
                year = int(row["year"])
                val  = float(row["emissions_total"])
            except (ValueError, KeyError):
                continue
            points.append({"d": str(year), "v": round(val / 1e9, 3)})

    points.sort(key=lambda p: p["d"])
    if len(points) < 10:
        print(f"  emissions: only {len(points)} points — skipping")
        return
    _write("emissions.json", {"id": "emissions", "unit": "Gt CO\u2082 / yr", "points": points})
    print(f"  emissions: {len(points)} annual points")


# ── Arctic sea ice (NSIDC G02135 v4) ─────────────────────────────────────────
# File: seaice_monyear_G02135_v4.0.xlsx  sheet: NH-Extent
# Wide format: row[0] = year, row[1..12] = Jan–Dec extent in M km²
# Output: { id, unit, points: [{d:"YYYY-MM", v:Mkm2}] }

def process_seaice():
    wb = openpyxl.load_workbook(
        DATA / "seaice_monyear_G02135_v4.0.xlsx",
        read_only=True, data_only=True
    )
    ws = wb["NH-Extent"]
    points = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        year = row[0]
        if not isinstance(year, (int, float)):
            continue
        year = int(year)
        for mi in range(12):
            v = row[mi + 1]  # columns 1–12 = Jan–Dec
            if v is None:
                continue
            points.append({"d": f"{year}-{mi + 1:02d}", "v": round(float(v), 3)})
    wb.close()
    points.sort(key=lambda p: p["d"])

    if not points:
        print("  seaice: no points — skipping")
        return
    _write("seaice.json", {"id": "seaice", "unit": "M km\u00b2", "points": points})
    print(f"  seaice: {len(points)} monthly points")


# ── US electricity-sector emissions by fuel (EIA) ────────────────────────────
# File: emissions_by_ind.xlsx  sheet: State Emissions
# Columns: Year, State, Producer Type, Energy Source, CO2 (Metric Tons), SO2, NOx
# Filters to Producer Type == "Total Electric Power Industry".
# Aggregates CO2 across all states, grouped by Year + mapped fuel key.
# Fuel map: Coal → coal, Natural Gas → gas, Petroleum → petroleum, rest → other
# Output: { id, unit, fuels, points: [{d:"YYYY", coal, gas, petroleum, other}] }

FUEL_MAP = {
    "Coal":                       "coal",
    "Natural Gas":                "gas",
    "Petroleum":                  "petroleum",
    "Other":                      "other",
    "Other Biomass":              "other",
    "Other Gases":                "other",
    "Geothermal":                 "other",
    "Wood and Wood Derived Fuels":"other",
}

def process_emissions_by_ind():
    wb = openpyxl.load_workbook(
        DATA / "emissions_by_ind.xlsx",
        read_only=True, data_only=True
    )
    ws = wb["State Emissions"]

    # year → fuel_key → cumulative CO2 (metric tons)
    totals = defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=2, values_only=True):
        year, _state, ptype, source, co2 = row[0], row[1], row[2], row[3], row[4]
        if ptype != "Total Electric Power Industry":
            continue
        if source == "All Sources":
            continue
        if not isinstance(year, int) or not isinstance(co2, (int, float)):
            continue
        key = FUEL_MAP.get(source)
        if key:
            totals[year][key] += co2
    wb.close()

    FUELS = ["coal", "gas", "petroleum", "other"]
    points = []
    for year in sorted(totals):
        pt = {"d": str(year)}
        for k in FUELS:
            pt[k] = round(totals[year].get(k, 0.0) / 1e6, 3)  # metric tons → Mt
        points.append(pt)

    if not points:
        print("  emissions_by_ind: no points — skipping")
        return
    _write("emissions_by_ind.json", {
        "id":     "emissions_by_ind",
        "unit":   "Mt CO\u2082",
        "fuels":  FUELS,
        "points": points,
    })
    print(f"  emissions_by_ind: {len(points)} annual points ({points[0]['d']}–{points[-1]['d']})")


# ── helpers ───────────────────────────────────────────────────────────────────

def _write(filename, data):
    path = OUT / filename
    path.write_text(json.dumps(data, separators=(",", ":")))


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Processing data → public/data/processed/")
    process_co2()
    process_gistemp()
    process_emissions()
    process_seaice()
    process_emissions_by_ind()
    print("Done.")
